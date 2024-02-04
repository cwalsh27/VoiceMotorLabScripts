import os
from openpyxl import *
del open

for filename in os.listdir():
    if filename.endswith(".xlsx"):
        # Load row numbers of selected trials
        wb = load_workbook(filename)
        ws = wb.active

        goodTrials = [(1, 0)]  # list of row numbers for button press trials, includes header row
        onsetTimes = []
        vowelCol = 0
        for col in range(1, 66):  # 66 is just three iterations of the alphabet, number is arbitrary
            if str(ws.cell(1, col).value).endswith('.RESP') and ('welcome' not in ws.cell(1, col).value.lower()):
                for row in range(2, 80):  # 80 to account for extra rows, 2 to skip line header
                    if str(ws.cell(row, col).value) != "None":
                        goodTrials.append((row, ws.cell(row, col+1).value))
            if str(ws.cell(1, col).value).endswith('.OnsetTime') and ('welcome' not in ws.cell(1, col).value.lower()):
                for row in range(1, 80):
                    onsetTimes.append(ws.cell(row, col).value)
            if ws.cell(1, col).value == 'vowels':
                vowelCol = col



        # new sheets for data of interest
        ws2 = wb.create_sheet('refined_data.xlsx')
        ws3 = wb.create_sheet('trial_start_times.xlsx')

        # define array
        goodVowels = []

        # populate refined_data
        row = 1
        if len(goodTrials) > 1:
            for goodTrial in goodTrials:
                for col in range(1, 66):
                    if (goodTrial[1] > 0) and (goodTrial[1] < 3000):
                        testVal = ws.cell(goodTrial[0]-1, col).value
                        ws2.cell(row, col).value = testVal
                if goodTrial[0] > 1:
                    goodVowels.append(ws.cell(goodTrial[0]-1, vowelCol).value)
                row += 1
        else:
            ws2.cell(1, 1).value = "No button presses during this trial"

        # populate trial_start_times
        ws3.cell(1, 1).value = str("Onset Time")
        #ws3.cell(1, 2).value = str("Pitch Presentation")
        #ws3.cell(1, 3).value = str("Metronome Onset")
        for i in range(2, len(onsetTimes)+1):
            if onsetTimes[i-1]:
                ws3.cell(i, 1).value = str(onsetTimes[i-1])
                #ws3.cell(i, 2).value = str(int(onsetTimes[i-1])+2400)
                #ws3.cell(i, 3).value = str(int(onsetTimes[i-1])+4000)


        # calculate vowel biases
        vowelString = ""
        for vowels in goodVowels:
            vowelString += vowels

        EaCount = vowelString.count("æ -ɛ - e - ɪ - i - ɪ- e - ɛ- æ")
        OuCount = vowelString.count("a - Ɔ - o - Ʊ – u – Ʊ – o – Ɔ – a")
        AaCount = vowelString.count("A / A")
        vowelCounts = [EaCount, OuCount, AaCount]
        '''
        row += 5
        ws2.cell(row, 1).value = "æ -ɛ - e - ɪ - i - ɪ- e - ɛ- æ"
        ws2.cell(row+1, 1).value = "a - Ɔ - o - Ʊ – u – Ʊ – o – Ɔ – a"
        ws2.cell(row+2, 1).value = "A / A"

        for i in range(0, 3):
            ws2.cell(row+i, 2).value = vowelCounts[i]
        '''
        # save spreadsheet changes
        outputFilename = "extracted" + str(filename)
        wb.save(outputFilename)


        # create/populate vowelBias text file
        vowelFile = open(filename[:-5] + "VowelBias.txt", "w")
        vowelFile.write("Ea Count: " + str(EaCount) + "\n")
        vowelFile.write("Ou Count: " + str(OuCount) + "\n")
        vowelFile.write("A/A Count: " + str(AaCount) + "\n")
        vowelFile.close()
